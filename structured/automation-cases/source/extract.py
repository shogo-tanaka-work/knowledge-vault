"""講師面談ログDB.xlsx を匿名化済み CSV に変換する。

入力: ~/Downloads/講師面談ログDB.xlsx（4シート）
出力: automation-cases.csv（全シート統合・氏名/社名マスク済み）

使い方:
    python extract.py [--xlsx PATH] [--out PATH]
"""
from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Iterable

import openpyxl

DEFAULT_XLSX = Path.home() / "Downloads" / "講師面談ログDB.xlsx"
SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_OUT = SCRIPT_DIR / "automation-cases.csv"
DICT_PATH = SCRIPT_DIR / "anonymize_dict.json"

UNIFIED_COLUMNS = [
    "id",
    "source_sheet",
    "source_row",
    "posted_at",
    "category",
    "tools",
    "outcome",
    "before",
    "after",
    "quant_effect",
    "qualitative_effect",
    "internal_rollout",
    "notes",
    "needs_industry_generalization",
]

# 各シートの列名 → 統一スキーマへのマッピング
SHEET_MAPPINGS: dict[str, dict[str, str]] = {
    "業務系の面談ログ_Gen": {
        "投稿日時": "posted_at",
        "業務課題カテゴリ": "category",
        "使用AIツール": "tools",
        "改善の成否": "outcome",
        "改善効果_定量": "quant_effect",
        "改善効果_定性": "qualitative_effect",
        "特記事項": "notes",
    },
    "業務系の面談ログ_Biz": {
        "投稿日時": "posted_at",
        "業務課題カテゴリ": "category",
        "使用AIツール": "tools",
        "改善の成否": "outcome",
        "改善効果_定量": "quant_effect",
        "改善効果_定性": "qualitative_effect",
        "特記事項": "notes",
    },
    "業務成果報告DB": {
        "投稿日時": "posted_at",
        "使用AIツール": "tools",
        "業務課題カテゴリ": "category",
        "改善前の課題": "before",
        "改善後の変化": "after",
        "改善効果_定量": "quant_effect",
        "改善効果_定性": "qualitative_effect",
        "社内展開の有無": "internal_rollout",
        "特記事項": "notes",
    },
    "Biz成果報告DB": {
        "投稿日時": "posted_at",
        "使用AIツール": "tools",
        "業務課題カテゴリ": "category",
        "改善前の課題": "before",
        "改善後の変化": "after",
        "改善効果_定量": "quant_effect",
        "改善効果_定性": "qualitative_effect",
        "社内展開の有無": "internal_rollout",
        "特記事項": "notes",
    },
}

# CSV から完全に除外する列（個人名）
DROP_COLUMNS = {"受講生名", "担当講師名", "投稿者", "message_id", "チャンネル"}

# 業界名が含まれそうなキーワード（マスク対象ではなくマーク用）
INDUSTRY_HINT_PATTERNS = [
    r"放送業界", r"放送局", r"映像制作", r"建設業", r"製造業", r"金融業",
    r"小売業", r"医療業界", r"教育業界", r"不動産", r"物流業",
]


def load_anonymize_dict() -> dict:
    if DICT_PATH.exists():
        return json.loads(DICT_PATH.read_text(encoding="utf-8"))
    return {"company_names": [], "person_names": [], "extra_terms": []}


def collect_person_names(wb: openpyxl.Workbook) -> set[str]:
    """xlsx 内に登場する受講生名・講師名を収集（マスク対象の基礎リスト）。"""
    names: set[str] = set()
    for sn in wb.sheetnames:
        ws = wb[sn]
        headers = [c.value for c in ws[1]]
        for col_name in ("受講生名", "担当講師名", "投稿者"):
            if col_name in headers:
                idx = headers.index(col_name)
                for row in ws.iter_rows(min_row=2, values_only=True):
                    v = row[idx]
                    if v and isinstance(v, str):
                        v = v.strip()
                        # 「(WAKO GROUP)」のような社名併記を分離
                        m = re.match(r"^([^（(]+)[（(]([^）)]+)[）)]\s*$", v)
                        if m:
                            names.add(m.group(1).strip())
                        else:
                            names.add(v)
    return {n for n in names if n}


def build_mask_patterns(person_names: set[str], anon_dict: dict) -> list[tuple[re.Pattern, str]]:
    patterns: list[tuple[re.Pattern, str]] = []

    # 個人名（長い順に置換）
    for name in sorted(person_names | set(anon_dict.get("person_names", [])), key=len, reverse=True):
        if len(name) >= 2:
            patterns.append((re.compile(re.escape(name)), "[氏名]"))

    # 社名（辞書由来）
    for company in sorted(anon_dict.get("company_names", []), key=len, reverse=True):
        patterns.append((re.compile(re.escape(company)), "[社名]"))

    # 法人格パターン
    patterns.append((re.compile(r"(?:株式会社|有限会社|合同会社|\(株\)|（株）)\s*[\w぀-ヿ一-鿿A-Za-z]+"), "[社名]"))
    patterns.append((re.compile(r"[\w一-鿿A-Za-z]+(?:株式会社|有限会社|合同会社|\(株\)|（株）)"), "[社名]"))

    # 追加用語
    for term in anon_dict.get("extra_terms", []):
        patterns.append((re.compile(re.escape(term)), "[マスク]"))

    return patterns


def mask(text: str, patterns: list[tuple[re.Pattern, str]]) -> tuple[str, bool]:
    """テキストにマスクを適用。業界名ヒントが残っていれば warn フラグを立てる。"""
    if not isinstance(text, str) or not text:
        return ("" if text is None else str(text), False)
    out = text
    for pat, repl in patterns:
        out = pat.sub(repl, out)
    needs_generalization = any(re.search(p, out) for p in INDUSTRY_HINT_PATTERNS)
    return out, needs_generalization


def extract(xlsx_path: Path, out_csv: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    anon_dict = load_anonymize_dict()
    person_names = collect_person_names(wb)
    patterns = build_mask_patterns(person_names, anon_dict)

    rows: list[dict] = []
    counts: dict[str, int] = {}

    for sheet_name, col_map in SHEET_MAPPINGS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        counts[sheet_name] = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            record: dict = {col: "" for col in UNIFIED_COLUMNS}
            record["source_sheet"] = sheet_name
            record["source_row"] = row_idx
            any_value = False
            any_warn = False

            for header, value in zip(headers, row):
                if header in DROP_COLUMNS:
                    continue
                target = col_map.get(header)
                if target is None:
                    continue
                if value is None:
                    continue
                if isinstance(value, str):
                    masked, warn = mask(value, patterns)
                    record[target] = masked
                    any_warn = any_warn or warn
                else:
                    record[target] = value
                any_value = True

            if not any_value:
                continue

            record["needs_industry_generalization"] = "Y" if any_warn else ""
            rows.append(record)
            counts[sheet_name] += 1

    # ID 付与（シート順 + 行順）
    for i, r in enumerate(rows, start=1):
        r["id"] = f"{i:04d}"

    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=UNIFIED_COLUMNS)
        writer.writeheader()
        for r in rows:
            writer.writerow(r)

    # 残存チェック
    leftover = []
    sample_text = out_csv.read_text(encoding="utf-8")
    for name in person_names:
        if len(name) >= 2 and name in sample_text:
            leftover.append(name)

    return {
        "out_csv": str(out_csv),
        "counts": counts,
        "total": len(rows),
        "person_names_collected": len(person_names),
        "leftover_person_names": leftover,
    }


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = ap.parse_args()

    if not args.xlsx.exists():
        raise SystemExit(f"xlsx not found: {args.xlsx}")

    result = extract(args.xlsx, args.out)
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
